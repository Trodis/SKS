# -*- coding: utf-8 -*-
import re
import click
import os
import sys
from bs4 import BeautifulSoup
import requests
import datetime
from requests.utils import quote
from ignoreconstants import ignore_openpyxl_constants                                               
ignore_openpyxl_constants()
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Style, Font
from sys import platform as _platform
if _platform == "win32":
    os.environ["REQUESTS_CA_BUNDLE"] = os.path.join(os.getcwd(), "cacert.pem")

now = datetime.datetime.now()
EXCELFILE = 'songkick_event_%s.xlsx' %now.strftime("%Y-%m-%d")

def parse_foundlocations(locations_html):
    bs_obj = BeautifulSoup(locations_html, "html.parser")
    locations_table = bs_obj.find('div', class_ = 'component filter-metro-area search')
    locations = {}
    counter = 1
    for l in locations_table.findAll('li'):
        for link in l.findAll('p'):
            locations.update({counter:[link.a.string, link.a['href']]})
            counter += 1
    return locations

def request_location(location):
    payload = {'utf8':'E29C93', 'query':location}
    result = requests.get('https://www.songkick.com/session/filter_metro_area?', params=payload)
    return result

def choose_locations(locations):
    print "\nLocations found on Songkick"
    print "*****************************"
    for k, v in locations.items():
        print "%s: %s" %(k, v[0])
    print "*****************************"

    while True:
        k = raw_input("Enter the Number of the Location: ")
        if k.isdigit():
            if int(k) in range(1, len(locations)+1):
                return locations[int(k)]
            else:
                continue

def request_events(location, payload=None):
    url = "https://www.songkick.com%s" %location[1]
    if not payload:
        return requests.get(url)
    else:
        return requests.get(url, params=payload)

def request_concertpage(concert_link):
    url = "https://songkick.com%s" %concert_link
    concert_page = requests.get(url)
    return concert_page

def parse_concertpage(concert_page, artist_name):
    bs_obj = BeautifulSoup(concert_page.text, "html.parser")
    tickets_list = []
    tickets_table = bs_obj.find('div', id='tickets')
    for tickets in tickets_table.find_all('div', class_='ticket-wrapper'):
        ticket_wrapper = tickets.find_all('div')
        if ticket_wrapper[1].span is not None:
            vendor = re.sub(r'\n\s*\n', r'\n\n',
                    ticket_wrapper[0].span.text.strip(), flags=re.M)
            ticket_price = re.sub(r'\n\s*\n', r'\n\n',
                    ticket_wrapper[1].span.text.strip(), flags=re.M) 
            ticket_link = tickets.a['href']
            tickets_list.append({'vendor':vendor, 'price': ticket_price, 'link': ticket_link})
        else:
            continue

    address = []
    venue_hcard = bs_obj.find('p', class_='venue-hcard')
    if venue_hcard is not None:
        for a in venue_hcard.find_all('span'):
            address.append(a.text)
    else:
        address.append("N/A")

    additional_details = bs_obj.find('div', class_='additional-details-container')
    if additional_details is not None:
        additional_details = additional_details.p.text
    else:
        additional_details = False

    date = bs_obj.find('div', class_='date-and-name').text.strip()
    city = bs_obj.find('div', class_='location').find_all('span')[1].text.strip()
    return tickets_list, address, city, additional_details, date

def write_sheet(artist_name, tickets, address, city, details, date, url, wb, ws):
    row_number = ws.max_row+1
    temp = row_number
    ws.cell(row=row_number, column=1, value=artist_name.string)
    ws.cell(row=row_number, column=2, value=url.strip())
    ws.cell(row=row_number, column=3, value=date.strip())
    for ticket in tickets:
        ticket_link = requests.head('https://www.songkick.com%s' %ticket['link'],
                allow_redirects=True).url 
        ws.cell(row=row_number, column=4, value=ticket['vendor'].strip())
        ws.cell(row=row_number, column=5, value=ticket_link.strip())
        ws.cell(row=row_number, column=6, value=ticket['price'])
        row_number = row_number+1
    row_number = temp
    ws.cell(row=row_number, column=7, value=city)
    ws.cell(row=row_number, column=8, value=','.join(address).strip())
    if details:
        ws.cell(row=row_number, column=9, value=details.strip())
    else:
        ws.cell(row=row_number, column=9, value='No Additional Details')
    wb.save(EXCELFILE) 

def create_sheet():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value='Name').font = Font(bold=True)
    ws.cell(row=1, column=2, value='Songkick URL').font = Font(bold=True)
    ws.cell(row=1, column=3, value='Date Start').font = Font(bold=True)
    ws.cell(row=1, column=4, value='Ticket Vendor').font = Font(bold=True)
    ws.cell(row=1, column=5, value='Ticket Link').font = Font(bold=True)
    ws.cell(row=1, column=6, value='Price').font = Font(bold=True)
    ws.cell(row=1, column=7, value='City').font = Font(bold=True)
    ws.cell(row=1, column=8, value='Address').font = Font(bold=True)
    ws.cell(row=1, column=9, value='Additional Details').font = Font(bold=True)
    wb.save(EXCELFILE)

    return wb, ws 

def parse_events(event_page, total_events, total_scraped, page_counter, wb, ws):
    bs_obj = BeautifulSoup(event_page, "html.parser")
    event_listings = bs_obj.find('ul', class_='event-listings ').find_all('li', class_=False)
    scraped = 0
    with click.progressbar(event_listings, label='(%s/%s) Scraping Page %s'\
            %(total_scraped, total_events, page_counter), length=len(event_listings),
            show_eta=False) as bar:
        for event in bar:
            sys.stdout.flush()
            if event.span.strong and event.a['href']:
                artist_name = event.span.strong
                link = event.a['href']
                concert_page = request_concertpage(link)
                tickets, address, city, details, date = parse_concertpage(concert_page, artist_name)
                url = concert_page.url
                write_sheet(artist_name, tickets, address, city, details, date, url, wb, ws)
                scraped = scraped + 1
            else:
                continue
        next_page = bs_obj.find('div', class_='pagination').find('a', class_='next_page')
    if next_page is not None:
        return next_page['href'], scraped
    else:
        return None, scraped

def main():
    stdout_encoding = sys.stdout.encoding
    wb, ws = create_sheet()
    location_entered_by_user = raw_input("Pleaser enter location: ")
    location_entered_by_user = location_entered_by_user.decode(stdout_encoding)

    location_html = request_location(location_entered_by_user).text
    locations = parse_foundlocations(location_html)
    selected_location = choose_locations(locations)

    initial_page = request_concertpage(selected_location[1])
    bs_obj = BeautifulSoup(initial_page.text, "html.parser")
    total_events = int(bs_obj.find('div', class_='component events-summary').h2.string.split(' ')[0])
    next_page = selected_location[1]
    page_counter = 1
    total_scraped = 0
    print ":: Scraping %s Events..." %total_events
    stop_scraping = False
    while next_page is not None:
        concert_page = request_concertpage(next_page)
        next_page, scraped = parse_events(concert_page.text, total_events, total_scraped,
                page_counter, wb, ws)
        total_scraped = total_scraped + scraped
        page_counter = page_counter + 1 
    print ":: Finished Scraping! ::"

if __name__ == '__main__':
    main()
